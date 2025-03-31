import openpyxl as xl
import os

FOLDER = r'G:\Config Management\As-Built\Parts Lists'

contents = os.listdir(FOLDER)

for item in contents:
    
    try:
        file_path = os.path.join(FOLDER,item)

        wb = xl.load_workbook(file_path)

        print(wb.sheetnames)
        del wb['UAB Parts Only']
        wb['Full'].title = 'Sheet 1'
        print(wb.sheetnames)

        wb.save(file_path)
    except:
        print('failed: ', item)